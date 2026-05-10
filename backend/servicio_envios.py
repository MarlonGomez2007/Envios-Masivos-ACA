import pandas as pd
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
import os
import time
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email.utils import make_msgid
import re
import socket
import json
from .gestion_cuentas import obtener_cuenta_disponible, incrementar_contador

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def conectar_servidor(servidor_anterior=None):
    if servidor_anterior:
        try:
            servidor_anterior.quit()
        except Exception:
            pass
    
    cuenta = obtener_cuenta_disponible()
    if not cuenta:
        return None, None
        
    try:
        servidor = smtplib.SMTP(cuenta["server"], cuenta["port"], timeout=60)
        servidor.ehlo()
        servidor.starttls()
        servidor.ehlo()
        servidor.login(cuenta["user"], cuenta["pass"])
        return servidor, cuenta
    except Exception as e:
        mensaje = str(e)
        if "Username and Password not accepted" in mensaje or "BadCredentials" in mensaje or "Invalid credentials" in mensaje:
            print(f"Error de credenciales SMTP para {cuenta['user']}: {mensaje}")
        else:
            print(f"Error conectando con {cuenta['user']}: {mensaje}")
        return None, cuenta

def enviar_correos(ruta_excel, pausa_segundos=5, modo_prueba=False, ruta_salida_excel=None, asunto="Envio Masivo", html="", firma_html="", imagenes_cids=None, adjuntos_paths=None, config=None, smtp=None, report=None):
    df = pd.read_excel(ruta_excel)
    df.columns = df.columns.str.strip()
    id_campana = str(int(time.time()))
    
    servidor = None
    cuenta_actual = None
    
    if not modo_prueba:
        servidor, cuenta_actual = conectar_servidor()
    
    resultados = []
    if imagenes_cids is None:
        imagenes_cids = {}
    if adjuntos_paths is None:
        adjuntos_paths = []
    tam_lote = 50
    pausa_lote = 30
    total_items = len(df)
    contador = 0
    
    for _, fila in df.iterrows():
        contador += 1
        nombre = ""
        apellido = ""
        correo = ""
        enlaces = {}
        if config and config.get("modo_nombre") == "completo":
            nombre = fila.get(config.get("col_nombre_completo") or "Nombre", "")
            apellido = ""
        else:
            nombre = fila.get(config.get("col_nombres") or "Nombres", "")
            apellido = fila.get(config.get("col_apellidos") or "Apellidos", "")
        correo = fila.get((config or {}).get("col_email") or "Email", "")
        cols_links = (config or {}).get("col_links") or ["link carton 1", "link carton 2"]
        for i, col in enumerate(cols_links, start=1):
            enlaces[f"link{i}"] = fila.get(col, "")
        plantilla = html or "<p>Hola {nombre}</p>"
        datos = {"nombre": nombre or "", "apellido": apellido or "", "email": correo or ""}
        for k, v in enlaces.items():
            datos[k] = v or ""
        mensaje_html = plantilla.format(**datos)
        if firma_html:
            mensaje_html = mensaje_html + firma_html
        for cid, ruta in imagenes_cids.items():
            mensaje_html = mensaje_html.replace(f"[imagen:{cid}]", f"<img src=\"cid:{cid}\">")
        mensaje = MIMEMultipart("mixed")
        relacionado = MIMEMultipart("related")
        
        remitente = cuenta_actual["user"] if cuenta_actual else "sin_remitente"
        mensaje["From"] = remitente
        mensaje["To"] = correo
        mensaje["Subject"] = asunto
        mensaje["Message-ID"] = make_msgid()
        parte_alternativa = MIMEMultipart("alternative")
        relacionado.attach(parte_alternativa)
        parte_alternativa.attach(MIMEText(mensaje_html, "html"))
        for cid, ruta in imagenes_cids.items():
            if os.path.exists(ruta):
                with open(ruta, "rb") as f:
                    img = MIMEImage(f.read())
                    img.add_header("Content-ID", f"<{cid}>")
                    relacionado.attach(img)
        mensaje.attach(relacionado)
        for ap in adjuntos_paths:
            if os.path.exists(ap):
                base = MIMEBase("application", "octet-stream")
                with open(ap, "rb") as f:
                    base.set_payload(f.read())
                encoders.encode_base64(base)
                nombre_adj = os.path.basename(ap)
                base.add_header("Content-Disposition", f"attachment; filename=\"{nombre_adj}\"")
                mensaje.attach(base)
        estado = ""
        codigo_error = ""
        descripcion_error = ""
        intentos = 0
        exito = False
        tipo_error = ""
        def dominio_resoluble(d):
            try:
                socket.getaddrinfo(d, 0)
                return True
            except Exception:
                return False
        dom = str(correo).split("@")[-1] if "@" in str(correo) else ""
        dominio_ok = dom and (dominio_resoluble(dom) or dominio_resoluble("mail."+dom) or dominio_resoluble("mx."+dom))
        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", str(correo)) or not dominio_ok:
            estado = "invalido"
            resultados.append({
                "Nombres": nombre,
                "Apellidos": apellido,
                "Email": correo,
                "Estado": estado,
                "Codigo": codigo_error,
                "Error": descripcion_error,
                "Fecha": pd.Timestamp.now(),
                "MessageID": mensaje["Message-ID"],
                "Campana": id_campana,
                "TiempoMs": 0,
                "Reintentos": intentos,
                "Tipo": tipo_error
            })
            if report:
                try:
                    report(correo, estado)
                except Exception:
                    pass
            continue
        if modo_prueba:
            resultados.append({
                "Nombres": nombre,
                "Apellidos": apellido,
                "Email": correo,
                "Estado": "prueba",
                "Codigo": "",
                "Error": "",
                "Fecha": pd.Timestamp.now(),
                "MessageID": mensaje["Message-ID"],
                "Campana": id_campana,
                "TiempoMs": 0,
                "Reintentos": 0,
                "Tipo": ""
            })
            time.sleep(pausa_segundos)
            if report:
                try:
                    report(correo, "prueba")
                except Exception:
                    pass
            if contador % tam_lote == 0 and contador < total_items:
                time.sleep(pausa_lote)
            continue
            
        if not servidor:
            servidor, cuenta_actual = conectar_servidor()
            
        if cuenta_actual and cuenta_actual.get("daily_count", 0) >= 450:
            servidor, cuenta_actual = conectar_servidor(servidor)
             
        if not servidor:
            estado = "fallo"
            if not cuenta_actual:
                descripcion_error = "No hay cuentas configuradas o todas superan el límite diario."
            else:
                descripcion_error = "Credenciales SMTP incorrectas o acceso bloqueado por el proveedor de correo."
            tipo_error = "configuracion"
        else:
            mensaje.replace_header("From", cuenta_actual["user"])
            while intentos < 3 and not exito:
                try:
                    t0 = time.time()
                    resp = servidor.sendmail(cuenta_actual["user"], correo, mensaje.as_string())
                    if isinstance(resp, dict) and resp:
                        estado = "rebotado"
                        datos = resp.get(correo) or next(iter(resp.values()))
                        codigo_error = datos[0] if isinstance(datos, tuple) and len(datos) > 0 else ""
                        descripcion_error = datos[1].decode("utf-8", "ignore") if isinstance(datos, tuple) and len(datos) > 1 and hasattr(datos[1], "decode") else str(datos)
                        tipo_error = "permanente"
                        exito = False
                        break
                    else:
                        estado = "enviado"
                        exito = True
                        incrementar_contador(cuenta_actual["user"])
                        cuenta_actual["daily_count"] += 1
                        cuenta_actual["last_sent_date"] = pd.Timestamp.now().isoformat()
                    tiempo_ms = int((time.time() - t0) * 1000)
                except smtplib.SMTPRecipientsRefused as e:
                    estado = "rebotado"
                    datos = list(e.recipients.values())[0] if getattr(e, "recipients", None) else None
                    codigo_error = datos[0] if datos else ""
                    descripcion_error = str(e)
                    tipo_error = "permanente"
                    break
                except (smtplib.SMTPSenderRefused, smtplib.SMTPDataError) as e:
                    estado = "fallo"
                    codigo_error = getattr(e, "smtp_code", None) or getattr(e, "code", None) or ""
                    descripcion_error = getattr(e, "smtp_error", None) or getattr(e, "message", None) or str(e)
                    intentos += 1
                    tipo_error = "permanente" if str(codigo_error).startswith("55") else "temporal"
                    if intentos < 3:
                        espera = min(8, 2 ** intentos)
                        time.sleep(espera)
                    else:
                        break
                except smtplib.SMTPServerDisconnected as e:
                    estado = "fallo"
                    codigo_error = ""
                    descripcion_error = str(e)
                    tipo_error = "temporal"
                    intentos += 1
                    servidor, cuenta_actual = conectar_servidor(servidor)
                    if not servidor:
                         break
                    espera = min(8, 2 ** intentos)
                    time.sleep(espera)
                except Exception as e:
                    estado = "fallo"
                    codigo_error = ""
                    descripcion_error = str(e)
                    tipo_error = "permanente"
                    break
        if exito and "tiempo_ms" not in locals():
            tiempo_ms = 0
        resultados.append({
            "Nombres": nombre,
            "Apellidos": apellido,
            "Email": correo,
            "Estado": estado,
            "Codigo": codigo_error,
            "Error": descripcion_error,
            "Fecha": pd.Timestamp.now(),
            "MessageID": mensaje["Message-ID"],
            "Campana": id_campana,
            "TiempoMs": tiempo_ms if "tiempo_ms" in locals() else 0,
            "Reintentos": intentos,
            "Tipo": tipo_error,
            "EnviadoDesde": cuenta_actual["user"] if cuenta_actual else ""
        })
        if report:
            try:
                report(correo, estado, cuenta_actual)
            except Exception:
                pass
        time.sleep(pausa_segundos)
        if contador % tam_lote == 0 and contador < total_items:
            time.sleep(pausa_lote)
    df_resultados = pd.DataFrame(resultados)
    if ruta_salida_excel:
        ruta_excel_salida = ruta_salida_excel
    else:
        ruta_excel_salida = os.path.join(BASE_DIR, "resultado_envios.xlsx")
    with pd.ExcelWriter(ruta_excel_salida, engine="openpyxl") as escritor:
        df_resultados.to_excel(escritor, index=False, sheet_name="Resultados")
        ws = escritor.book.active
        fuente_encabezado = Font(bold=True, color="FFFFFF")
        relleno_encabezado = PatternFill("solid", fgColor="4F81BD")
        alineacion_centro = Alignment(horizontal="center", vertical="center")
        borde = Border(left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"), top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))
        for c in range(1, ws.max_column + 1):
            celda = ws.cell(row=1, column=c)
            celda.font = fuente_encabezado
            celda.fill = relleno_encabezado
            celda.alignment = alineacion_centro
            celda.border = borde
        for r in range(2, ws.max_row + 1):
            color_fila = PatternFill("solid", fgColor="F7F7F7") if r % 2 == 0 else None
            for c in range(1, ws.max_column + 1):
                celda = ws.cell(row=r, column=c)
                celda.border = borde
                if color_fila:
                    celda.fill = color_fila
        for i, nombre_col in enumerate(df_resultados.columns, start=1):
            max_largo = max([len(str(nombre_col))] + [len(str(x)) if x is not None else 0 for x in df_resultados[nombre_col].tolist()])
            ancho = min(max_largo + 6, 80)
            ws.column_dimensions[get_column_letter(i)].width = ancho
        if "Error" in df_resultados.columns:
            idx_error = df_resultados.columns.get_loc("Error") + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx_error).alignment = Alignment(wrap_text=True, vertical="top")
        if "Fecha" in df_resultados.columns:
            idx_fecha = df_resultados.columns.get_loc("Fecha") + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx_fecha).number_format = "yyyy-mm-dd hh:mm"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws2 = escritor.book.create_sheet("Resumen")
        total = len(df_resultados)
        enviados = int((df_resultados["Estado"] == "enviado").sum())
        rebotados = int((df_resultados["Estado"] == "rebotado").sum())
        fallos = int((df_resultados["Estado"] == "fallo").sum())
        invalidos = int((df_resultados["Estado"] == "invalido").sum())
        pruebas = int((df_resultados["Estado"] == "prueba").sum())
        tasa_envio = enviados / total if total else 0
        tasa_rebote = rebotados / total if total else 0
        tasa_fallo = fallos / total if total else 0
        tasa_invalido = invalidos / total if total else 0
        datos_resumen = [
            ["Total", total],
            ["Enviados", enviados],
            ["Rebotados", rebotados],
            ["Fallos", fallos],
            ["Inválidos", invalidos],
            ["Pruebas", pruebas],
            ["Tasa Envío", tasa_envio],
            ["Tasa Rebote", tasa_rebote],
            ["Tasa Fallo", tasa_fallo],
            ["Tasa Inválido", tasa_invalido]
        ]
        ws2.append(["Métrica", "Valor"])
        for fila in datos_resumen:
            ws2.append(fila)
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 18
        for c in range(1, 3):
            celda = ws2.cell(row=1, column=c)
            celda.font = fuente_encabezado
            celda.fill = relleno_encabezado
            celda.alignment = alineacion_centro
            celda.border = borde
        for r in range(2, ws2.max_row + 1):
            color_fila = PatternFill("solid", fgColor="F7F7F7") if r % 2 == 0 else None
            for c in range(1, 3):
                celda = ws2.cell(row=r, column=c)
                celda.border = borde
                if color_fila:
                    celda.fill = color_fila
        for r in range(2, ws2.max_row + 1):
            ws2.cell(row=r, column=2).number_format = "0.00%" if r >= 7 else "0"
    if servidor:
        try:
            servidor.quit()
        except Exception:
            pass
    return ruta_excel_salida
