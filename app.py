from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, flash
import os
import shutil
import tempfile
import zipfile
import barcode
from barcode.writer import ImageWriter
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Alignment, Font, Border, Side
from PIL import Image as PilImage, ImageDraw, ImageFont
from io import BytesIO
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from backend.servicio_envios import enviar_correos
from backend.gestion_cuentas import cargar_cuentas, agregar_o_actualizar_cuenta, eliminar_cuenta
from backend.auth import obtener_usuario, verificar_credenciales, crear_usuario, eliminar_usuario, cambiar_password, obtener_todos_usuarios, obtener_historial, registrar_accion, cambiar_estado_usuario, Usuario
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from drive_enlaces_a_excel import obtener_id_carpeta, autenticar, listar_archivos, guardar_excel, obtener_nombre_carpeta
from backend.cartones_manager import obtener_cartones_usados, registrar_envios, obtener_reporte_evento, extraer_numero_carton
import pandas as pd
import json
from datetime import datetime
from googleapiclient.discovery import build
import logging
import threading

ruta_base = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__)
app.secret_key = "super_secret_key_premium_app_v1"

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

@login_manager.user_loader
def load_user(user_id):
    return obtener_usuario(user_id)
logger = logging.getLogger("envio_masivos")
logger.setLevel(logging.ERROR)
if not logger.handlers:
    ruta_log = os.path.join(ruta_base, "salidas", "app_error.log")
    manejador = logging.FileHandler(ruta_log, encoding="utf-8")
    manejador.setLevel(logging.ERROR)
    formato = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    manejador.setFormatter(formato)
    logger.addHandler(manejador)
app.config["UPLOAD_FOLDER"] = os.path.join(ruta_base, "uploads")
app.config["OUTPUT_FOLDER"] = os.path.join(ruta_base, "salidas")
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["OUTPUT_FOLDER"], exist_ok=True)
PROGRESOS = {}
RUTA_API_KEY_RAIZ = os.path.join(ruta_base, "api_key.txt")
RUTA_API_KEY_SALIDAS = os.path.join(app.config["OUTPUT_FOLDER"], "api_key.txt")
SMTP_CONF_PATH = os.path.join(ruta_base, "smtp_config.txt")


@app.errorhandler(Exception)
def manejar_error(e):
    logger.exception("Error no controlado")
    return "Error interno de servidor", 500

if os.path.exists(SMTP_CONF_PATH) and not cargar_cuentas():
    try:
        with open(SMTP_CONF_PATH, "r", encoding="utf-8") as f:
            content = f.read().strip()
            try:
                d = json.loads(content)
                if d.get("user"):
                    agregar_o_actualizar_cuenta({
                        "server": d.get("server"),
                        "port": d.get("port"),
                        "user": d.get("user"),
                        "pass": d.get("pass")
                    })
            except json.JSONDecodeError:
                parts = content.split(",")
                if len(parts) >= 4:
                    agregar_o_actualizar_cuenta({
                        "server": parts[0].strip(),
                        "port": int(parts[1].strip()),
                        "user": parts[2].strip(),
                        "pass": parts[3].strip()
                    })
    except Exception:
        pass

def leer_api_key():
    for ruta in (RUTA_API_KEY_SALIDAS, RUTA_API_KEY_RAIZ):
        if os.path.exists(ruta):
            try:
                with open(ruta, "r", encoding="utf-8") as f:
                    return f.read().strip()
            except Exception:
                continue
    return ""

def guardar_api_key(k):
    valor = str(k or "").strip()
    for ruta in (RUTA_API_KEY_SALIDAS, RUTA_API_KEY_RAIZ):
        try:
            with open(ruta, "w", encoding="utf-8") as f:
                f.write(valor)
        except Exception:
            continue

def validar_base_excel(ruta, cfg):
    try:
        xls = pd.ExcelFile(ruta)
        if "Base" not in xls.sheet_names:
            return False
        df = pd.read_excel(ruta, sheet_name="Base", nrows=0)
        esperados = []
        if cfg.get("modo_nombre") == "completo":
            esperados.append(cfg.get("col_nombre_completo") or "Nombre")
        else:
            esperados.extend([cfg.get("col_nombres") or "Nombres", cfg.get("col_apellidos") or "Apellidos"])
        esperados.append(cfg.get("col_email") or "Email")
        esperados.extend(cfg.get("col_links") or [])
        reales = [str(x).strip() for x in df.columns.tolist()]
        esperados = [str(x).strip() for x in esperados]
        return all(col in reales for col in esperados)
    except Exception:
        return False

def validar_base_excel_detalle(ruta, cfg):
    try:
        xls = pd.ExcelFile(ruta)
        if "Base" not in xls.sheet_names:
            return False, [], []
        df = pd.read_excel(ruta, sheet_name="Base", nrows=0)
        esperados = []
        if cfg.get("modo_nombre") == "completo":
            esperados.append(cfg.get("col_nombre_completo") or "Nombre")
        else:
            esperados.extend([cfg.get("col_nombres") or "Nombres", cfg.get("col_apellidos") or "Apellidos"])
        esperados.append(cfg.get("col_email") or "Email")
        esperados.extend(cfg.get("col_links") or [])
        reales = [str(x).strip() for x in df.columns.tolist()]
        esperados = [str(x).strip() for x in esperados]
        ok = all(col in reales for col in esperados)
        return ok, esperados, reales
    except Exception:
        return False, [], []

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("inicio"))
    
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        user, msg = verificar_credenciales(username, password)
        
        if user:
            login_user(user)
            registrar_accion(user.id, "inicio_sesion")
            next_page = request.args.get("next")
            return redirect(next_page or url_for("inicio"))
        else:
            if msg == "Usuario desactivado":
                flash("Su usuario se encuentra desactivado", "error")
            else:
                flash("Usuario o contraseña incorrectos", "error")
            
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    registrar_accion(current_user.id, "cierre_sesion")
    logout_user()
    return redirect(url_for("login"))

@app.route("/admin", methods=["GET", "POST"])
@login_required
def admin_panel():
    if not current_user.is_admin:
        return redirect(url_for("inicio"))
    
    if request.method == "POST":
        accion = request.form.get("accion")
        if accion == "crear":
            u = request.form.get("username")
            p = request.form.get("password")
            r = request.form.get("role") or "user"
            ok, msg = crear_usuario(u, p, r)
            flash(msg, "success" if ok else "error")
            if ok: registrar_accion(current_user.id, "crear_usuario", {"target": u, "role": r})
            
        elif accion == "eliminar":
            u = request.form.get("username")
            if u != current_user.id:
                eliminar_usuario(u)
                flash("Usuario eliminado", "success")
                registrar_accion(current_user.id, "eliminar_usuario", {"target": u})
            else:
                flash("No puedes eliminarte a ti mismo", "error")
                
        elif accion == "password":
            u = request.form.get("username")
            p = request.form.get("password")
            if cambiar_password(u, p):
                flash("Contraseña actualizada", "success")
                registrar_accion(current_user.id, "cambiar_password", {"target": u})
            else:
                flash("Error al cambiar contraseña", "error")

        elif accion == "estado":
            u = request.form.get("username")
            estado = request.form.get("active") == "true"
            if u != current_user.id:
                if cambiar_estado_usuario(u, estado):
                    flash(f"Usuario {'activado' if estado else 'inactivado'} correctamente", "success")
                    registrar_accion(current_user.id, "cambio_estado", {"target": u, "nuevo_estado": estado})
                else:
                    flash("Error al cambiar estado", "error")
            else:
                flash("No puedes desactivar tu propia cuenta", "error")
                
        return redirect(url_for("admin_panel"))

    usuarios = obtener_todos_usuarios()
    historial = obtener_historial()
    return render_template("admin.html", usuarios=usuarios, historial=historial)

@app.route("/")
@login_required
def inicio():
    return render_template("index.html")

@app.route("/enviar", methods=["POST"])
@login_required
def enviar():
    pausa = 5.0
    prueba = request.form.get("prueba") == "on"
    asunto = request.form.get("asunto") or "Envio Masivo"
    contenido = request.form.get("html") or ""
    firma = request.form.get("firma") or ""
    modo = request.form.get("modo_nombre") or "completo"
    col_nc = request.form.get("col_nombre_completo") or "Nombre"
    col_n = request.form.get("col_nombres") or "Nombres"
    col_a = request.form.get("col_apellidos") or "Apellidos"
    col_e = request.form.get("col_email") or "Email"
    cant_links = int(request.form.get("links_cantidad") or "1")
    cols_links = request.form.getlist("col_links") or []
    if not cols_links or len(cols_links) < cant_links:
        cols_links = [f"Link {i}" for i in range(1, cant_links + 1)]
    cfg = {
        "modo_nombre": modo,
        "col_nombre_completo": col_nc,
        "col_nombres": col_n,
        "col_apellidos": col_a,
        "col_email": col_e,
        "col_links": cols_links
    }
    archivo = request.files.get("archivo")
    if not archivo:
        return redirect(url_for("inicio"))
    nombre_archivo = secure_filename(archivo.filename)
    ruta_archivo = os.path.join(app.config["UPLOAD_FOLDER"], nombre_archivo)
    archivo.save(ruta_archivo)
    if not validar_base_excel(ruta_archivo, cfg):
        return redirect(url_for("inicio") + "?error_estructura=1")
    imagenes = request.files.getlist("imagenes")
    adjuntos = request.files.getlist("adjuntos")
    cids = {}
    adj_paths = []
    for img in imagenes:
        if not img.filename:
            continue
        n = secure_filename(img.filename)
        p = os.path.join(app.config["UPLOAD_FOLDER"], n)
        img.save(p)
        cids[n] = p
    for a in adjuntos:
        if not a.filename:
            continue
        n = secure_filename(a.filename)
        p = os.path.join(app.config["UPLOAD_FOLDER"], n)
        a.save(p)
        adj_paths.append(p)
        if n.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')):
            cids[n] = p
    ruta_salida = os.path.join(app.config["OUTPUT_FOLDER"], "resultado_envios.xlsx")
    registrar_accion(current_user.id, "envio_simple", {"archivo": nombre_archivo, "asunto": asunto})
    
    salida = enviar_correos(ruta_archivo, pausa, prueba, ruta_salida, asunto, contenido, firma, cids, adj_paths, cfg)
    return redirect(url_for("resultado"))

@app.route("/resultado")
@login_required
def resultado():
    ruta_salida = os.path.join(app.config["OUTPUT_FOLDER"], "resultado_envios.xlsx")
    if os.path.exists(ruta_salida):
        return render_template("resultado.html")
    return redirect(url_for("inicio"))

@app.route("/descargar")
@login_required
def descargar():
    ruta_salida = os.path.join(app.config["OUTPUT_FOLDER"], "resultado_envios.xlsx")
    return send_file(ruta_salida, as_attachment=True)

@app.route("/admin/descargar-zip/<filename>")
@login_required
def descargar_zip_generado(filename):
    if not current_user.is_admin:
        return redirect(url_for("inicio"))
    filename = secure_filename(filename)
    ruta_zip = os.path.join(app.config["OUTPUT_FOLDER"], filename)
    
    if os.path.exists(ruta_zip):
        return send_file(ruta_zip, as_attachment=True)
    return "Archivo no encontrado", 404

@app.route("/admin/generar-codigos", methods=["GET", "POST"])
@login_required
def generar_codigos():
    if not current_user.is_admin:
        return redirect(url_for("inicio"))
        
    if request.method == "POST":
        try:
            v_ini_str = request.form.get("visual_inicio")
            v_fin_str = request.form.get("visual_fin")
            
            if not v_ini_str or not v_fin_str:
                return jsonify({"ok": False, "msg": "Debe ingresar el rango visual"})

            v_ini = int(v_ini_str)
            v_fin = int(v_fin_str)
            usar_diff = request.form.get("usar_diferente") == "on"
            
            r_ini = int(request.form.get("real_inicio") or v_ini)
            r_fin = int(request.form.get("real_fin") or v_fin)
            
            usar_guia = request.form.get("usar_guia") == "on"
            g_ini = 0
            g_fin = 0
            
            if usar_guia:
                g_ini = int(request.form.get("guia_inicio") or v_ini)
                g_fin = int(request.form.get("guia_fin") or v_fin)
            
            if v_fin < v_ini:
                return jsonify({"ok": False, "msg": "El rango visual es inválido (Final menor que Inicio)"})
                
            if usar_diff:
                if (v_fin - v_ini) != (r_fin - r_ini):
                    return jsonify({"ok": False, "msg": "Los rangos deben tener la misma cantidad de elementos"})
            
            if usar_guia:
                if (v_fin - v_ini) != (g_fin - g_ini):
                    return jsonify({"ok": False, "msg": "El rango guía debe tener la misma cantidad de elementos"})
            
            with tempfile.TemporaryDirectory() as temp_dir:
                imgs_dir = os.path.join(temp_dir, "imagenes")
                os.makedirs(imgs_dir)
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Codigos"
                
                wb_nums = Workbook()
                ws_nums = wb_nums.active
                ws_nums.title = "Numeros"
                columnas = ['A', 'B', 'C', 'D']
                for col in columnas:
                    ws.column_dimensions[col].width = 28
                    ws_nums.column_dimensions[col].width = 28
                
                count = (v_fin - v_ini) + 1
                
                borde_delgado = Side(border_style="thin", color="CCCCCC")
                borde_corte = Border(left=borde_delgado, right=borde_delgado, top=borde_delgado, bottom=borde_delgado)
                
                for i in range(count):
                    val_visual = v_ini + i
                    val_real = r_ini + i if usar_diff else val_visual
                    val_guia = g_ini + i if usar_guia else None
                    
                    code_val = str(val_real)
                    filename_base = f"{val_visual}"
                    filename = os.path.join(imgs_dir, filename_base)
                    writer = ImageWriter()
                    rv = barcode.get('code128', code_val, writer=writer)
                    options = {
                        'write_text': True,
                        'module_height': 15.0,
                        'module_width': 0.45,
                        'quiet_zone': 4.0,
                        'font_size': 14,
                        'text_distance': 6.0,
                        'background': 'white',
                        'foreground': 'black'
                    }
                    saved_path = rv.save(filename, options)
                    
                    try:
                        with PilImage.open(saved_path) as im:
                            if im.mode != 'RGB':
                                im = im.convert('RGB')
                            
                            from PIL import ImageChops
                            bg = PilImage.new(im.mode, im.size, im.getpixel((0,0)))
                            diff = ImageChops.difference(im, bg)
                            diff = ImageChops.add(diff, diff, 2.0, -100)
                            bbox = diff.getbbox()
                            
                            if bbox:
                                im_cropped = im.crop(bbox)
                            else:
                                im_cropped = im

                            im_cropped.save(saved_path)
                            
                    except Exception as e:
                        logger.error(f"Error procesando imagen {filename}: {e}")

                    img = ExcelImage(saved_path)
                    target_height = 110
                    aspect_ratio = img.width / img.height
                    target_width = int(target_height * aspect_ratio)
                    img.width = target_width
                    img.height = target_height
                    c = i % 4
                    row_base = (i // 4) + 1
                    r_img = row_base
                    col_letter = columnas[c]
                    cell_img = ws[f"{col_letter}{r_img}"]
                    cell_img.border = Border(left=borde_delgado, right=borde_delgado, top=borde_delgado, bottom=borde_delgado)
                    cell_img.alignment = Alignment(horizontal='center', vertical='center')
                    if usar_guia and val_guia is not None:
                        cell_img.value = val_guia
                        cell_img.alignment = Alignment(horizontal='left', vertical='bottom')
                        cell_img.font = Font(size=10, color="000000", bold=True)
                    else:
                        cell_img.value = ""
                    CELL_W_PX = 212
                    CELL_H_PX = 127
                    
                    off_x = max(0, (CELL_W_PX - target_width) // 2)
                    off_y = max(0, (CELL_H_PX - target_height) // 2)
                    
                    marker = AnchorMarker(col=c, colOff=pixels_to_EMU(off_x), row=r_img-1, rowOff=pixels_to_EMU(off_y))
                    size = XDRPositiveSize2D(pixels_to_EMU(target_width), pixels_to_EMU(target_height))
                    img.anchor = OneCellAnchor(_from=marker, ext=size)
                    ws.add_image(img)
                    cell_num = ws_nums[f"{col_letter}{r_img}"]
                    cell_num.border = Border(left=borde_delgado, right=borde_delgado, top=borde_delgado, bottom=borde_delgado)
                    if usar_guia and val_guia is not None:
                        cell_num.value = val_guia
                        cell_num.alignment = Alignment(horizontal='left', vertical='bottom')
                        cell_num.font = Font(size=10, color="000000", bold=True)
                    
                    try:
                        img_w, img_h = 400, 200
                        txt_img = PilImage.new('RGBA', (img_w, img_h), (255, 255, 255, 0))
                        d = ImageDraw.Draw(txt_img)
                        try:
                            fnt = ImageFont.truetype("arialbd.ttf", 140)
                        except:
                            try:
                                fnt = ImageFont.truetype("arial.ttf", 140)
                            except:
                                fnt = ImageFont.load_default()
                        bbox = d.textbbox((0, 0), code_val, font=fnt)
                        text_w = bbox[2] - bbox[0]
                        text_h = bbox[3] - bbox[1]
                        pos_x = (img_w - text_w) / 2
                        pos_y = (img_h - text_h) / 2
                        d.text((pos_x, pos_y), code_val, font=fnt, fill=(0, 0, 0, 255))
                        bbox_full = txt_img.getbbox()
                        if bbox_full:
                            txt_img = txt_img.crop(bbox_full)
                        num_filename = f"num_{val_visual}.png"
                        num_path = os.path.join(imgs_dir, num_filename)
                        txt_img.save(num_path, "PNG")
                        img_num_xl = ExcelImage(num_path)
                        num_w, num_h = txt_img.size
                        max_w = 175
                        max_h = 100
                        ratio = min(max_w/num_w, max_h/num_h) if num_w > 0 and num_h > 0 else 1
                        target_w_num = int(num_w * ratio)
                        target_h_num = int(num_h * ratio)
                        img_num_xl.width = target_w_num
                        img_num_xl.height = target_h_num
                        off_x_num = max(0, ((CELL_W_PX - target_w_num) // 2) - 10)
                        off_y_num = max(0, (CELL_H_PX - target_h_num) // 2)
                        
                        marker_num = AnchorMarker(col=c, colOff=pixels_to_EMU(off_x_num), row=r_img-1, rowOff=pixels_to_EMU(off_y_num))
                        size_num = XDRPositiveSize2D(pixels_to_EMU(target_w_num), pixels_to_EMU(target_h_num))
                        img_num_xl.anchor = OneCellAnchor(_from=marker_num, ext=size_num)
                        
                        ws_nums.add_image(img_num_xl)
                        
                    except Exception as e:
                        logger.error(f"Error generando imagen de numero: {e}")
                        cell_num.value = code_val
                        cell_num.alignment = Alignment(horizontal='center', vertical='center')
                        cell_num.font = Font(size=70, bold=True)
                    
                    ws.row_dimensions[r_img].height = 95
                    ws_nums.row_dimensions[r_img].height = 95

                excel_path = os.path.join(temp_dir, "Grilla_Codigos.xlsx")
                wb.save(excel_path)
                excel_nums_path = os.path.join(temp_dir, "Grilla_Numeros.xlsx")
                wb_nums.save(excel_nums_path)
                folder_name = f"Codigos_{v_ini}-{v_fin}"
                zip_filename = f"Codigos_{v_ini}-{v_fin}.zip"
                zip_path = os.path.join(app.config["OUTPUT_FOLDER"], zip_filename)
                
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    zf.write(excel_path, f"{folder_name}/Grilla_Codigos.xlsx")
                    zf.write(excel_nums_path, f"{folder_name}/Grilla_Numeros.xlsx")
                    for root, dirs, files in os.walk(imgs_dir):
                        for file in files:
                            zf.write(os.path.join(root, file), f"{folder_name}/Imagenes/{file}")
                            
            registrar_accion(current_user.id, "generar_codigos", {"rango": f"{v_ini}-{v_fin}"})
            return jsonify({
                "ok": True, 
                "url": url_for('descargar_zip_generado', filename=zip_filename)
            })
            
        except Exception as e:
            logger.exception("Error generando códigos")
            return jsonify({"ok": False, "msg": f"Error interno: {str(e)}"})
            
    return render_template("generar_codigos.html")


def _generar_codigos_desde_lista(codigos_visuales, codigos_reales, codigos_guia, usar_guia):
    """Genera los dos Excel + imágenes desde listas de valores. Devuelve ruta del ZIP."""
    with tempfile.TemporaryDirectory() as temp_dir:
        imgs_dir = os.path.join(temp_dir, "imagenes")
        os.makedirs(imgs_dir)

        wb = Workbook()
        ws = wb.active
        ws.title = "Codigos"
        wb_nums = Workbook()
        ws_nums = wb_nums.active
        ws_nums.title = "Numeros"

        columnas = ['A', 'B', 'C', 'D']
        for col in columnas:
            ws.column_dimensions[col].width = 28
            ws_nums.column_dimensions[col].width = 28

        borde_delgado = Side(border_style="thin", color="CCCCCC")
        CELL_W_PX = 212
        CELL_H_PX = 127

        for i, val_visual in enumerate(codigos_visuales):
            val_real = codigos_reales[i]
            val_guia = codigos_guia[i] if usar_guia else None

            code_val = str(val_real)
            filename_base = f"{val_visual}"
            filename = os.path.join(imgs_dir, filename_base)

            writer = ImageWriter()
            rv = barcode.get('code128', code_val, writer=writer)
            options = {
                'write_text': True,
                'module_height': 15.0,
                'module_width': 0.45,
                'quiet_zone': 4.0,
                'font_size': 14,
                'text_distance': 6.0,
                'background': 'white',
                'foreground': 'black'
            }
            saved_path = rv.save(filename, options)

            try:
                with PilImage.open(saved_path) as im:
                    if im.mode != 'RGB':
                        im = im.convert('RGB')
                    from PIL import ImageChops
                    bg = PilImage.new(im.mode, im.size, im.getpixel((0, 0)))
                    diff = ImageChops.difference(im, bg)
                    diff = ImageChops.add(diff, diff, 2.0, -100)
                    bbox = diff.getbbox()
                    im_cropped = im.crop(bbox) if bbox else im
                    im_cropped.save(saved_path)
            except Exception as e:
                logger.error(f"Error procesando imagen {filename}: {e}")

            img = ExcelImage(saved_path)
            target_height = 110
            aspect_ratio = img.width / img.height
            target_width = int(target_height * aspect_ratio)
            img.width = target_width
            img.height = target_height

            c = i % 4
            row_base = (i // 4) + 1
            r_img = row_base
            col_letter = columnas[c]

            cell_img = ws[f"{col_letter}{r_img}"]
            cell_img.border = Border(left=borde_delgado, right=borde_delgado, top=borde_delgado, bottom=borde_delgado)

            if usar_guia and val_guia is not None:
                cell_img.value = val_guia
                cell_img.alignment = Alignment(horizontal='left', vertical='bottom')
                cell_img.font = Font(size=10, color="000000", bold=True)
            else:
                cell_img.value = ""
                cell_img.alignment = Alignment(horizontal='center', vertical='center')

            off_x = max(0, (CELL_W_PX - target_width) // 2)
            off_y = max(0, (CELL_H_PX - target_height) // 2)
            marker = AnchorMarker(col=c, colOff=pixels_to_EMU(off_x), row=r_img - 1, rowOff=pixels_to_EMU(off_y))
            size = XDRPositiveSize2D(pixels_to_EMU(target_width), pixels_to_EMU(target_height))
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)

            cell_num = ws_nums[f"{col_letter}{r_img}"]
            cell_num.border = Border(left=borde_delgado, right=borde_delgado, top=borde_delgado, bottom=borde_delgado)

            if usar_guia and val_guia is not None:
                cell_num.value = val_guia
                cell_num.alignment = Alignment(horizontal='left', vertical='bottom')
                cell_num.font = Font(size=10, color="000000", bold=True)

            try:
                img_w, img_h = 400, 200
                txt_img = PilImage.new('RGBA', (img_w, img_h), (255, 255, 255, 0))
                d = ImageDraw.Draw(txt_img)
                try:
                    fnt = ImageFont.truetype("arialbd.ttf", 140)
                except:
                    try:
                        fnt = ImageFont.truetype("arial.ttf", 140)
                    except:
                        fnt = ImageFont.load_default()

                bbox2 = d.textbbox((0, 0), code_val, font=fnt)
                text_w = bbox2[2] - bbox2[0]
                text_h = bbox2[3] - bbox2[1]
                pos_x = (img_w - text_w) / 2
                pos_y = (img_h - text_h) / 2
                d.text((pos_x, pos_y), code_val, font=fnt, fill=(0, 0, 0, 255))
                bbox_full = txt_img.getbbox()
                if bbox_full:
                    txt_img = txt_img.crop(bbox_full)

                num_filename = f"num_{val_visual}.png"
                num_path = os.path.join(imgs_dir, num_filename)
                txt_img.save(num_path, "PNG")

                img_num_xl = ExcelImage(num_path)
                num_w, num_h = txt_img.size
                max_w, max_h = 175, 100
                ratio = min(max_w / num_w, max_h / num_h) if num_w > 0 and num_h > 0 else 1
                target_w_num = int(num_w * ratio)
                target_h_num = int(num_h * ratio)
                img_num_xl.width = target_w_num
                img_num_xl.height = target_h_num

                off_x_num = max(0, ((CELL_W_PX - target_w_num) // 2) - 10)
                off_y_num = max(0, (CELL_H_PX - target_h_num) // 2)
                marker_num = AnchorMarker(col=c, colOff=pixels_to_EMU(off_x_num), row=r_img - 1, rowOff=pixels_to_EMU(off_y_num))
                size_num = XDRPositiveSize2D(pixels_to_EMU(target_w_num), pixels_to_EMU(target_h_num))
                img_num_xl.anchor = OneCellAnchor(_from=marker_num, ext=size_num)
                ws_nums.add_image(img_num_xl)

            except Exception as e:
                logger.error(f"Error generando imagen de numero: {e}")
                cell_num.value = code_val
                cell_num.alignment = Alignment(horizontal='center', vertical='center')
                cell_num.font = Font(size=70, bold=True)

            ws.row_dimensions[r_img].height = 95
            ws_nums.row_dimensions[r_img].height = 95

        excel_path = os.path.join(temp_dir, "Grilla_Codigos.xlsx")
        wb.save(excel_path)
        excel_nums_path = os.path.join(temp_dir, "Grilla_Numeros.xlsx")
        wb_nums.save(excel_nums_path)

        v_ini = codigos_visuales[0]
        v_fin = codigos_visuales[-1]
        folder_name = f"Codigos_{v_ini}-{v_fin}"
        zip_filename = f"Codigos_{v_ini}-{v_fin}.zip"
        zip_path = os.path.join(app.config["OUTPUT_FOLDER"], zip_filename)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(excel_path, f"{folder_name}/Grilla_Codigos.xlsx")
            zf.write(excel_nums_path, f"{folder_name}/Grilla_Numeros.xlsx")
            for root, dirs, files in os.walk(imgs_dir):
                for file in files:
                    zf.write(os.path.join(root, file), f"{folder_name}/Imagenes/{file}")

        return zip_filename


@app.route("/admin/generar-codigos-excel", methods=["POST"])
@login_required
def generar_codigos_excel():
    if not current_user.is_admin:
        return jsonify({"ok": False, "msg": "Sin permisos"})

    try:
        archivo = request.files.get("archivo_excel")
        if not archivo or archivo.filename == "":
            return jsonify({"ok": False, "msg": "No se subió ningún archivo"})

        usar_guia = request.form.get("usar_guia_excel") == "on"
        g_ini_str = request.form.get("guia_inicio_excel", "").strip()
        g_fin_str = request.form.get("guia_fin_excel", "").strip()

        df = pd.read_excel(archivo, header=0)
        col = df.columns[0]
        valores_raw = df[col].dropna().tolist()

        codigos_visuales = []
        for v in valores_raw:
            try:
                codigos_visuales.append(int(v))
            except (ValueError, TypeError):
                pass

        if not codigos_visuales:
            return jsonify({"ok": False, "msg": "No se encontraron códigos válidos en el archivo"})

        count = len(codigos_visuales)
        codigos_reales = list(codigos_visuales)

        codigos_guia = []
        if usar_guia:
            if not g_ini_str or not g_fin_str:
                return jsonify({"ok": False, "msg": "Debe ingresar el rango guía"})
            g_ini = int(g_ini_str)
            g_fin = int(g_fin_str)
            if (g_fin - g_ini + 1) != count:
                return jsonify({"ok": False, "msg": f"El rango guía debe tener {count} elementos, tiene {g_fin - g_ini + 1}"})
            codigos_guia = list(range(g_ini, g_fin + 1))

        zip_filename = _generar_codigos_desde_lista(codigos_visuales, codigos_reales, codigos_guia, usar_guia)

        registrar_accion(current_user.id, "generar_codigos_excel", {"count": count})

        return jsonify({
            "ok": True,
            "url": url_for('descargar_zip_generado', filename=zip_filename)
        })

    except Exception as e:
        logger.exception("Error generando códigos desde Excel")
        return jsonify({"ok": False, "msg": f"Error interno: {str(e)}"})


@app.context_processor
def inject_flags():
    ruta = os.path.join(app.config["OUTPUT_FOLDER"], "resultado_envios.xlsx")
    return {"has_resultado": os.path.exists(ruta)}

@app.route("/descargar-plantilla", methods=["POST"])
@login_required
def descargar_plantilla():
    modo = request.form.get("modo_nombre") or "completo"
    col_nc = request.form.get("col_nombre_completo") or "Nombre"
    col_n = request.form.get("col_nombres") or "Nombres"
    col_a = request.form.get("col_apellidos") or "Apellidos"
    col_e = request.form.get("col_email") or "Email"
    cant_links = int(request.form.get("links_cantidad") or "1")
    cols_links = request.form.getlist("col_links") or [f"Link {i}" for i in range(1, cant_links + 1)]
    wb = Workbook()
    ws = wb.active
    ws.title = "Base"
    encabezados = []
    if modo == "completo":
        encabezados.append(col_nc)
    else:
        encabezados.extend([col_n, col_a])
    encabezados.append(col_e)
    encabezados.extend(cols_links[:cant_links])
    ws.append(encabezados)
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    fuente = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="5EC9FF")
    borde = Border(left=Side(style="thin", color="9DE2FF"), right=Side(style="thin", color="9DE2FF"), top=Side(style="thin", color="9DE2FF"), bottom=Side(style="thin", color="9DE2FF"))
    for i, h in enumerate(encabezados, start=1):
        c = ws.cell(row=1, column=i)
        c.font = fuente
        c.fill = fondo
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde
        col = get_column_letter(i)
        hl = str(h).lower()
        ancho = 28
        if "email" in hl:
            ancho = 42
        elif hl.startswith("link"):
            ancho = 60
        elif "nombre" in hl or "apellido" in hl:
            ancho = 34
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = "A2"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="plantilla_envios.xlsx")
@app.route("/validar-base", methods=["POST"])
@login_required
def validar_base():
    modo = request.form.get("modo_nombre") or "completo"
    col_nc = request.form.get("col_nombre_completo") or "Nombre"
    col_n = request.form.get("col_nombres") or "Nombres"
    col_a = request.form.get("col_apellidos") or "Apellidos"
    col_e = request.form.get("col_email") or "Email"
    cant_links = int(request.form.get("links_cantidad") or "1")
    cols_links = request.form.getlist("col_links") or []
    if not cols_links or len(cols_links) < cant_links:
        cols_links = [f"Link {i}" for i in range(1, cant_links + 1)]
    cfg = {
        "modo_nombre": modo,
        "col_nombre_completo": col_nc,
        "col_nombres": col_n,
        "col_apellidos": col_a,
        "col_email": col_e,
        "col_links": cols_links
    }
    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename:
        return jsonify({"ok": False, "esperados": [], "reales": [], "filas": []})
    nombre_archivo = secure_filename(archivo.filename)
    ruta_archivo = os.path.join(app.config["UPLOAD_FOLDER"], nombre_archivo)
    archivo.save(ruta_archivo)
    ok, esperados, reales = validar_base_excel_detalle(ruta_archivo, cfg)
    filas = []
    try:
        df = pd.read_excel(ruta_archivo, sheet_name="Base")
        df.columns = df.columns.str.strip()
        max_prev = 1000
        cuenta = 0
        for _, fila in df.iterrows():
            if cuenta >= max_prev:
                break
            nombre_mostrar = ""
            if cfg.get("modo_nombre") == "completo":
                nombre_mostrar = str(fila.get(cfg.get("col_nombre_completo") or "Nombre") or "")
            else:
                nn = str(fila.get(cfg.get("col_nombres") or "Nombres") or "")
                aa = str(fila.get(cfg.get("col_apellidos") or "Apellidos") or "")
                nombre_mostrar = (nn + " " + aa).strip()
            email_mostrar = str(fila.get(cfg.get("col_email") or "Email") or "")
            enlaces = []
            for col in cfg.get("col_links") or []:
                enlaces.append(str(fila.get(col) or ""))
            filas.append({"nombre": nombre_mostrar, "email": email_mostrar, "links": enlaces})
            cuenta += 1
    except Exception:
        filas = []
    try:
        os.remove(ruta_archivo)
    except Exception:
        pass
    return jsonify({"ok": ok, "esperados": esperados, "reales": reales, "filas": filas})

@app.route("/enviar-json", methods=["POST"])
@login_required
def enviar_json():
    pausa = 5.0
    prueba = request.form.get("prueba") == "on"
    asunto = request.form.get("asunto") or "Envio Masivo"
    contenido = request.form.get("html") or ""
    firma = request.form.get("firma") or ""
    modo = request.form.get("modo_nombre") or "completo"
    col_nc = request.form.get("col_nombre_completo") or "Nombre"
    col_n = request.form.get("col_nombres") or "Nombres"
    col_a = request.form.get("col_apellidos") or "Apellidos"
    col_e = request.form.get("col_email") or "Email"
    cant_links = int(request.form.get("links_cantidad") or "1")
    cols_links = request.form.getlist("col_links") or []
    if not cols_links or len(cols_links) < cant_links:
        cols_links = [f"Link {i}" for i in range(1, cant_links + 1)]
    cfg = {
        "modo_nombre": modo,
        "col_nombre_completo": col_nc,
        "col_nombres": col_n,
        "col_apellidos": col_a,
        "col_email": col_e,
        "col_links": cols_links
    }
    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename:
        return jsonify({"ok": False, "resultados": []})
    nombre_archivo = secure_filename(archivo.filename)
    ruta_archivo = os.path.join(app.config["UPLOAD_FOLDER"], nombre_archivo)
    archivo.save(ruta_archivo)
    if not validar_base_excel(ruta_archivo, cfg):
        try:
            os.remove(ruta_archivo)
        except Exception:
            pass
        return jsonify({"ok": False, "resultados": []})
    ruta_salida = os.path.join(app.config["OUTPUT_FOLDER"], "resultado_envios.xlsx")
    
    registrar_accion(current_user.id, "envio_json", {"archivo": nombre_archivo, "asunto": asunto})
    enviar_correos(ruta_archivo, pausa, prueba, ruta_salida, asunto, contenido, firma, {}, [], cfg)
    resultados = []
    try:
        dfres = pd.read_excel(ruta_salida, sheet_name="Resultados")
        for _, fila in dfres.iterrows():
            email = str(fila.get("Email") or "")
            estado = str(fila.get("Estado") or "")
            resultados.append({"email": email, "estado": estado})
    except Exception:
        resultados = []
    try:
        os.remove(ruta_archivo)
    except Exception:
        pass
    return jsonify({"ok": True, "resultados": resultados})

@app.route("/enviar-iniciar", methods=["POST"])
@login_required
def enviar_iniciar():
    pausa = 5.0
    prueba = request.form.get("prueba") == "on"
    asunto = request.form.get("asunto") or "Envio Masivo"
    contenido = request.form.get("html") or ""
    firma = request.form.get("firma") or ""
    modo = request.form.get("modo_nombre") or "completo"
    col_nc = request.form.get("col_nombre_completo") or "Nombre"
    col_n = request.form.get("col_nombres") or "Nombres"
    col_a = request.form.get("col_apellidos") or "Apellidos"
    col_e = request.form.get("col_email") or "Email"
    cant_links = int(request.form.get("links_cantidad") or "1")
    cols_links = request.form.getlist("col_links") or []
    if not cols_links or len(cols_links) < cant_links:
        cols_links = [f"Link {i}" for i in range(1, cant_links + 1)]
    cfg = {
        "modo_nombre": modo,
        "col_nombre_completo": col_nc,
        "col_nombres": col_n,
        "col_apellidos": col_a,
        "col_email": col_e,
        "col_links": cols_links
    }
    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename:
        return jsonify({"ok": False})
    nombre_archivo = secure_filename(archivo.filename)
    ruta_archivo = os.path.join(app.config["UPLOAD_FOLDER"], nombre_archivo)
    archivo.save(ruta_archivo)
    imagenes = request.files.getlist("imagenes")
    adjuntos = request.files.getlist("adjuntos")
    cids = {}
    adj_paths = []
    for img in imagenes:
        if not img.filename:
            continue
        n = secure_filename(img.filename)
        p = os.path.join(app.config["UPLOAD_FOLDER"], n)
        img.save(p)
        cids[n] = p
    for a in adjuntos:
        if not a.filename:
            continue
        n = secure_filename(a.filename)
        p = os.path.join(app.config["UPLOAD_FOLDER"], n)
        a.save(p)
        adj_paths.append(p)
        if n.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')):
            cids[n] = p
    ruta_salida = os.path.join(app.config["OUTPUT_FOLDER"], "resultado_envios.xlsx")
    registrar_accion(current_user.id, "envio_background", {"archivo": nombre_archivo, "asunto": asunto})
    
    proc_id = str(int(pd.Timestamp.now().timestamp()))
    PROGRESOS[proc_id] = {"items": [], "done": False, "status": {}}
    def report(email, estado, cuenta=None):
        p = PROGRESOS.get(proc_id, {})
        p.get("items", []).append({"email": str(email or ""), "estado": str(estado or "")})
        if cuenta:
            last_date = cuenta.get("last_sent_date", "")
            formatted_date = ""
            if last_date:
                try:
                    dt = datetime.fromisoformat(last_date)
                    meses = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
                    mes = meses[dt.month - 1]
                    hora = dt.strftime("%I:%M %p")
                    formatted_date = f"{dt.day} {mes} {dt.year}, {hora}"
                except:
                    formatted_date = last_date
            
            p["status"] = {
                "cuenta_user": cuenta.get("user", ""),
                "cuenta_daily": cuenta.get("daily_count", 0),
                "cuenta_limit": 450,
                "email_actual": str(email or ""),
                "last_sent_date": formatted_date
            }
    def tarea():
        enviar_correos(ruta_archivo, pausa, prueba, ruta_salida, asunto, contenido, firma, cids, adj_paths, cfg, report=report)
        PROGRESOS[proc_id]["done"] = True
        try:
            os.remove(ruta_archivo)
        except Exception:
            pass
    h = threading.Thread(target=tarea, daemon=True)
    h.start()
    return jsonify({"ok": True, "id": proc_id})

@app.route("/progreso", methods=["GET"])
@login_required
def progreso():
    proc_id = request.args.get("id") or ""
    off = int(request.args.get("offset") or "0")
    dat = PROGRESOS.get(proc_id) or {"items": [], "done": True}
    items = dat.get("items", [])
    done = dat.get("done", True)
    status = dat.get("status", {})
    nuevos = items[off:] if off < len(items) else []
    return jsonify({"ok": True, "items": nuevos, "done": done, "total": len(items), "status": status})

@app.route("/configuracion", methods=["GET", "POST"])
@login_required
def configuracion():
    if request.method == "GET":
        cuentas = cargar_cuentas()
        meses = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
        for c in cuentas:
            last = c.get("last_sent_date")
            if last:
                try:
                    dt = datetime.fromisoformat(last)
                    mes = meses[dt.month - 1]
                    hora = dt.strftime("%I:%M %p")
                    c["last_sent_fmt"] = f"{dt.day} {mes} {dt.year}, {hora}"
                except:
                    c["last_sent_fmt"] = last
            else:
                c["last_sent_fmt"] = "Sin actividad reciente"
                
        return render_template("configuracion.html", cuentas=cuentas, api_key=leer_api_key())
    
    accion = request.form.get("accion")
    
    if accion == "eliminar":
        email_eliminar = request.form.get("email_eliminar")
        if email_eliminar:
            eliminar_cuenta(email_eliminar)
    else:
        server = request.form.get("server") or ""
        port = int(request.form.get("port") or "587")
        user = request.form.get("user") or ""
        pwd = request.form.get("pass") or ""
        if user and pwd:
            agregar_o_actualizar_cuenta({"server": server, "port": port, "user": user, "pass": pwd})
            
        api_key = request.form.get("api_key")
        if api_key is not None:
            guardar_api_key(api_key)
            
    return redirect(url_for("configuracion") + "?ok=1")

@app.route("/drive", methods=["GET", "POST"])
@login_required
def drive():
    if request.method == "GET":
        return render_template("drive.html")
    url = request.form.get("url")
    if not url:
        return redirect(url_for("drive"))
    carpeta_id = obtener_id_carpeta(url)
    creds = autenticar()
    servicio = None
    if creds:
        servicio = build("drive", "v3", credentials=creds)
    else:
        clave = leer_api_key()
        if clave:
            servicio = build("drive", "v3", developerKey=clave)
    if not servicio:
        return redirect(url_for("drive"))
    datos = listar_archivos(servicio, carpeta_id)
    nombre_carpeta = obtener_nombre_carpeta(servicio, carpeta_id)
    nombre_archivo = secure_filename(f"enlaces_{nombre_carpeta}.xlsx")
    ruta_excel = os.path.join(app.config["OUTPUT_FOLDER"], nombre_archivo)
    guardar_excel(datos, ruta_excel)
    return send_file(ruta_excel, as_attachment=True)

@app.route("/drive_info", methods=["POST"])
def drive_info():
    url = request.form.get("url")
    if not url:
        return jsonify({"ok": False})
    try:
        carpeta_id = obtener_id_carpeta(url)
    except Exception:
        return jsonify({"ok": False})
    creds = autenticar()
    servicio = None
    if creds:
        servicio = build("drive", "v3", credentials=creds)
    else:
        clave = leer_api_key()
        if clave:
            servicio = build("drive", "v3", developerKey=clave)
    if not servicio:
        return jsonify({"ok": False})
    try:
        datos = listar_archivos(servicio, carpeta_id)
        nombre_carpeta = obtener_nombre_carpeta(servicio, carpeta_id)
        return jsonify({"ok": True, "nombre": nombre_carpeta, "cantidad": len(datos)})
    except Exception:
        return jsonify({"ok": False})

@app.route("/drive_generar", methods=["POST"])
def drive_generar():
    url = request.form.get("url")
    if not url:
        return jsonify({"ok": False})
    try:
        carpeta_id = obtener_id_carpeta(url)
    except Exception:
        return jsonify({"ok": False})
    creds = autenticar()
    servicio = None
    if creds:
        servicio = build("drive", "v3", credentials=creds)
    else:
        clave = leer_api_key()
        if clave:
            servicio = build("drive", "v3", developerKey=clave)
    if not servicio:
        return jsonify({"ok": False})
    try:
        datos = listar_archivos(servicio, carpeta_id)
        nombre_carpeta = obtener_nombre_carpeta(servicio, carpeta_id)
        nombre_archivo = secure_filename(f"enlaces_{nombre_carpeta}.xlsx")
        ruta_excel = os.path.join(app.config["OUTPUT_FOLDER"], nombre_archivo)
        guardar_excel(datos, ruta_excel)
        tipos = {}
        total_tamano = 0
        ultima_fecha = None
        for d in datos:
            mt = str(d.get("mimeType") or "")
            if mt.startswith("image/"):
                k = "Imágenes"
            elif mt == "application/pdf":
                k = "PDF"
            elif mt.startswith("video/"):
                k = "Videos"
            elif mt.startswith("audio/"):
                k = "Audio"
            elif mt.startswith("text/"):
                k = "Texto"
            elif mt.startswith("application/vnd.google-apps"):
                k = "Google Docs"
            else:
                k = "Archivos"
            tipos[k] = tipos.get(k, 0) + 1
            try:
                total_tamano += int(d.get("size") or 0)
            except Exception:
                pass
            f = d.get("mod")
            if f and (not ultima_fecha or f > ultima_fecha):
                ultima_fecha = f
        import math
        def convert_size(size_bytes):
            if size_bytes == 0:
                return "0B"
            size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
            i = int(math.floor(math.log(size_bytes, 1024)))
            p = math.pow(1024, i)
            s = round(size_bytes / p, 2)
            return "%s %s" % (s, size_name[i])
        
        return jsonify({
            "ok": True,
            "nombre": nombre_carpeta,
            "cantidad": len(datos),
            "tipos": tipos,
            "tamano": total_tamano,
            "modificado": str(ultima_fecha or ""),
            "ultima_mod": str(ultima_fecha or "")[:10],
            "archivo": nombre_archivo,
            "descarga_url": url_for("drive_descargar")
        })
    except Exception:
        return jsonify({"ok": False})

@app.route("/drive_descargar")
def drive_descargar():
    files = [f for f in os.listdir(app.config["OUTPUT_FOLDER"]) if f.startswith("enlaces_") and f.endswith(".xlsx")]
    if not files:
        return redirect(url_for("drive"))
    files.sort(key=lambda x: os.path.getmtime(os.path.join(app.config["OUTPUT_FOLDER"], x)), reverse=True)
    ruta = os.path.join(app.config["OUTPUT_FOLDER"], files[0])
    return send_file(ruta, as_attachment=True)

@app.route("/cartones-preview-excel", methods=["POST"])
@login_required
def cartones_preview_excel():
    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename:
        return jsonify({"ok": False, "filas": []})
    nombre_archivo = secure_filename(archivo.filename)
    ruta_archivo = os.path.join(app.config["UPLOAD_FOLDER"], nombre_archivo)
    archivo.save(ruta_archivo)
    filas = []
    try:
        try:
            df = pd.read_excel(ruta_archivo, sheet_name="Base")
        except Exception:
            df = pd.read_excel(ruta_archivo)
        df.columns = df.columns.str.strip()
        cols_lower = {c.lower(): c for c in df.columns}
        col_nombre = None
        col_email = None
        col_cantidad = None
        for k, v in cols_lower.items():
            if "nombre" in k and col_nombre is None:
                col_nombre = v
            if "email" in k or "correo" in k:
                col_email = v
            if "cantidad" in k or "carton" in k or "cartón" in k:
                col_cantidad = v
        if not col_nombre:
            col_nombre = df.columns[0]
        for _, fila in df.iterrows():
            nombre = str(fila.get(col_nombre, "") or "").strip()
            email = str(fila.get(col_email, "") or "").strip() if col_email else ""
            try:
                cantidad = int(fila.get(col_cantidad, 1) or 1) if col_cantidad else 1
            except Exception:
                cantidad = 1
            filas.append({"nombre": nombre, "email": email, "cantidad": cantidad})
    except Exception as e:
        logger.error(f"Error leyendo excel preview cartones: {e}")
    try:
        os.remove(ruta_archivo)
    except Exception:
        pass
    return jsonify({"ok": True, "filas": filas})


@app.route("/envio-cartones/descargar-asignacion", methods=["POST"])
@login_required
def cartones_descargar_asignacion():
    """Genera Excel con la asignación simulada (sin enviar) para revisión previa."""
    drive_url = request.form.get("drive_url", "").strip()
    excepciones_json = request.form.get("excepciones", "[]")
    archivo = request.files.get("archivo")

    if not drive_url or not archivo or not archivo.filename:
        return "Faltan parámetros", 400

    try:
        excepciones = json.loads(excepciones_json)
    except Exception:
        excepciones = []

    nombre_archivo = secure_filename(archivo.filename)
    ruta_archivo = os.path.join(app.config["UPLOAD_FOLDER"], nombre_archivo)
    archivo.save(ruta_archivo)

    try:
        try:
            df = pd.read_excel(ruta_archivo, sheet_name="Base")
        except Exception:
            df = pd.read_excel(ruta_archivo)
        df.columns = df.columns.str.strip()
    except Exception as e:
        return f"Error leyendo Excel: {e}", 400
    finally:
        try:
            os.remove(ruta_archivo)
        except Exception:
            pass

    cols_lower = {c.lower(): c for c in df.columns}
    col_nombre = None
    col_email = None
    col_cantidad = None
    for k, v in cols_lower.items():
        if "nombre" in k and col_nombre is None:
            col_nombre = v
        if "email" in k or "correo" in k:
            col_email = v
        if "cantidad" in k or "carton" in k or "cartón" in k:
            col_cantidad = v
    if not col_nombre:
        col_nombre = df.columns[0]
    if not col_email or not col_cantidad:
        return "Columnas requeridas no encontradas", 400

    try:
        carpeta_id = obtener_id_carpeta(drive_url)
        servicio = _obtener_servicio_drive()
        if not servicio:
            return "No se pudo conectar a Drive", 500
        datos_drive = listar_archivos(servicio, carpeta_id)
    except Exception as e:
        return f"Error Drive: {e}", 500

    mapa_cartones = {}
    for d in datos_drive:
        num = extraer_numero_carton(d.get("nombre", "") or d.get("name", ""))
        if num is not None:
            mapa_cartones[num] = d.get("enlace", "") or d.get("link", "")

    cartones_ya_usados = obtener_cartones_usados(drive_url)

    mapa_excepciones = {}
    for exc in excepciones:
        email_exc = str(exc.get("email", "")).strip().lower()
        rini = int(exc.get("rango_ini", 0))
        rfin = int(exc.get("rango_fin", 0))
        if email_exc and rfin >= rini:
            mapa_excepciones[email_exc] = (rini, rfin)

    asignacion = _calcular_asignacion(df, col_nombre, col_email, col_cantidad, mapa_cartones, cartones_ya_usados, mapa_excepciones)

    from openpyxl import Workbook as WB
    from openpyxl.styles import Font as FT, PatternFill as PF, Alignment as AL, Border as BR, Side as SD
    from openpyxl.utils import get_column_letter

    wb = WB()
    ws = wb.active
    ws.title = "Asignación"

    encabezados = ["Nombre", "Email", "Cantidad", "Cartones asignados", "Números de cartón", "Nota"]
    ws.append(encabezados)

    fuente_enc = FT(bold=True, color="FFFFFF")
    relleno_enc = PF("solid", fgColor="4F81BD")
    borde = BR(left=SD(style="thin", color="BFBFBF"), right=SD(style="thin", color="BFBFBF"),
               top=SD(style="thin", color="BFBFBF"), bottom=SD(style="thin", color="BFBFBF"))
    al_c = AL(horizontal="center", vertical="center")
    al_l = AL(horizontal="left", vertical="top", wrap_text=True)

    for c in range(1, len(encabezados) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = fuente_enc
        cel.fill = relleno_enc
        cel.alignment = al_c
        cel.border = borde

    for i, item in enumerate(asignacion, start=2):
        nums = item["cartones"]
        links = item["links"]
        links_str = "\n".join(links) if links else ""
        nums_str = ", ".join(str(n) for n in nums) if nums else "—"
        fila_xl = [item["nombre"], item["email"], len(nums), links_str, nums_str, item.get("nota", "")]
        ws.append(fila_xl)
        color = "F7F9FC" if i % 2 == 0 else "FFFFFF"
        relleno_f = PF("solid", fgColor=color)
        for c in range(1, len(encabezados) + 1):
            cel = ws.cell(row=i, column=c)
            cel.border = borde
            cel.fill = relleno_f
            cel.alignment = al_l

    anchos = [30, 35, 12, 60, 25, 25]
    for c, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(c)].width = ancho

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="asignacion_cartones.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _obtener_servicio_drive():
    creds = autenticar()
    if creds:
        return build("drive", "v3", credentials=creds)
    clave = leer_api_key()
    if clave:
        return build("drive", "v3", developerKey=clave)
    return None


def _calcular_asignacion(df, col_nombre, col_email, col_cantidad, mapa_cartones, cartones_ya_usados, mapa_excepciones):
    nums_reservados = set()
    for rini, rfin in mapa_excepciones.values():
        for n in range(rini, rfin + 1):
            nums_reservados.add(n)

    pool_global = sorted(set(mapa_cartones.keys()) - cartones_ya_usados - nums_reservados)
    pool_idx = 0
    usados_sesion = set(cartones_ya_usados)
    resultado = []

    for _, fila in df.iterrows():
        nombre = str(fila.get(col_nombre, "") or "").strip()
        correo = str(fila.get(col_email, "") or "").strip()
        try:
            cantidad = int(fila.get(col_cantidad, 1) or 1)
        except Exception:
            cantidad = 1

        if not correo or "@" not in correo:
            resultado.append({"nombre": nombre, "email": correo, "cartones": [], "links": [], "nota": "email inválido"})
            continue

        correo_lower = correo.lower()
        cartones_asignados = []

        if correo_lower in mapa_excepciones:
            rini, rfin = mapa_excepciones[correo_lower]
            cartones_asignados = [n for n in range(rini, rfin + 1) if n in mapa_cartones and n not in usados_sesion]
        else:
            while len(cartones_asignados) < cantidad and pool_idx < len(pool_global):
                num = pool_global[pool_idx]
                pool_idx += 1
                if num not in usados_sesion:
                    cartones_asignados.append(num)

        for num in cartones_asignados:
            usados_sesion.add(num)

        links = [mapa_cartones.get(n, "") for n in cartones_asignados]
        nota = "excepción (rango fijo)" if correo_lower in mapa_excepciones else ""
        if not cartones_asignados:
            nota = "sin cartones disponibles"
        resultado.append({"nombre": nombre, "email": correo, "cartones": cartones_asignados, "links": links, "nota": nota})

    return resultado


@app.route("/envio-cartones", methods=["GET"])
@login_required
def envio_cartones():
    return render_template("envio_cartones.html")


@app.route("/envio-cartones/info-drive", methods=["POST"])
@login_required
def cartones_info_drive():
    url = request.form.get("url", "").strip()
    if not url:
        return jsonify({"ok": False, "msg": "URL requerida"})
    try:
        carpeta_id = obtener_id_carpeta(url)
        servicio = _obtener_servicio_drive()
        if not servicio:
            return jsonify({"ok": False, "msg": "No se pudo conectar a Drive"})
        datos = listar_archivos(servicio, carpeta_id)
        nombre_carpeta = obtener_nombre_carpeta(servicio, carpeta_id)
        cartones_disponibles = {}
        for d in datos:
            nombre = d.get("nombre", "") or d.get("name", "")
            enlace = d.get("enlace", "") or d.get("link", "")
            num = extraer_numero_carton(nombre)
            if num is not None:
                cartones_disponibles[num] = {"nombre": nombre, "link": enlace}
        usados = obtener_cartones_usados(url)
        nums_disponibles = sorted(set(cartones_disponibles.keys()) - usados)
        return jsonify({
            "ok": True,
            "nombre_carpeta": nombre_carpeta,
            "total_archivos": len(datos),
            "total_cartones": len(cartones_disponibles),
            "cartones_usados": len(usados),
            "cartones_usados_lista": sorted(usados),
            "cartones_disponibles": len(nums_disponibles),
            "rango_min": min(cartones_disponibles.keys()) if cartones_disponibles else 0,
            "rango_max": max(cartones_disponibles.keys()) if cartones_disponibles else 0,
        })
    except Exception as e:
        logger.exception("Error info drive cartones")
        return jsonify({"ok": False, "msg": str(e)})


@app.route("/envio-cartones/iniciar", methods=["POST"])
@login_required
def cartones_iniciar():
    drive_url = request.form.get("drive_url", "").strip()
    asunto = request.form.get("asunto", "Tu cartón de bingo")
    html_body = request.form.get("html", "")
    firma = request.form.get("firma", "")
    excepciones_json = request.form.get("excepciones", "[]")

    try:
        excepciones = json.loads(excepciones_json)
    except Exception:
        excepciones = []

    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename:
        return jsonify({"ok": False, "msg": "Se requiere el archivo Excel de destinatarios"})
    if not drive_url:
        return jsonify({"ok": False, "msg": "Se requiere la URL de Drive"})

    nombre_archivo = secure_filename(archivo.filename)
    ruta_archivo = os.path.join(app.config["UPLOAD_FOLDER"], nombre_archivo)
    archivo.save(ruta_archivo)

    try:
        try:
            df = pd.read_excel(ruta_archivo, sheet_name="Base")
        except Exception:
            df = pd.read_excel(ruta_archivo)
        df.columns = df.columns.str.strip()
    except Exception as e:
        return jsonify({"ok": False, "msg": f"Error leyendo Excel: {e}"})

    cols_lower = {c.lower(): c for c in df.columns}
    col_nombre = None
    col_email = None
    col_cantidad = None
    for k, v in cols_lower.items():
        if "nombre" in k and col_nombre is None:
            col_nombre = v
        if "email" in k or "correo" in k:
            col_email = v
        if "cantidad" in k or "carton" in k or "cartón" in k:
            col_cantidad = v

    if not col_email:
        return jsonify({"ok": False, "msg": "No se encontró columna de Email en el archivo"})
    if not col_nombre:
        col_nombre = df.columns[0]
    if not col_cantidad:
        return jsonify({"ok": False, "msg": "No se encontró columna de 'cantidad de cartones' en el archivo"})

    try:
        carpeta_id = obtener_id_carpeta(drive_url)
        servicio = _obtener_servicio_drive()
        if not servicio:
            return jsonify({"ok": False, "msg": "No se pudo conectar a Drive"})
        datos_drive = listar_archivos(servicio, carpeta_id)
    except Exception as e:
        return jsonify({"ok": False, "msg": f"Error conectando a Drive: {e}"})

    mapa_cartones = {}
    for d in datos_drive:
        nombre = d.get("nombre", "") or d.get("name", "")
        enlace = d.get("enlace", "") or d.get("link", "")
        num = extraer_numero_carton(nombre)
        if num is not None:
            mapa_cartones[num] = enlace

    if not mapa_cartones:
        return jsonify({"ok": False, "msg": "No se encontraron cartones en la carpeta de Drive"})

    cartones_ya_usados = obtener_cartones_usados(drive_url)

    mapa_excepciones = {}
    for exc in excepciones:
        email_exc = str(exc.get("email", "")).strip().lower()
        rini = int(exc.get("rango_ini", 0))
        rfin = int(exc.get("rango_fin", 0))
        if email_exc and rfin >= rini:
            mapa_excepciones[email_exc] = (rini, rfin)

    proc_id = str(int(pd.Timestamp.now().timestamp()))
    PROGRESOS[proc_id] = {"items": [], "done": False, "status": {}, "tipo": "cartones"}

    def tarea():
        from backend.servicio_envios import conectar_servidor
        from backend.gestion_cuentas import incrementar_contador

        nums_reservados = set()
        for rini, rfin in mapa_excepciones.values():
            for n in range(rini, rfin + 1):
                nums_reservados.add(n)

        pool_global = sorted(set(mapa_cartones.keys()) - cartones_ya_usados - nums_reservados)
        envios_realizados = []
        resumen_items = []
        servidor, cuenta_actual = conectar_servidor()
        pool_idx = 0

        for _, fila in df.iterrows():
            nombre = str(fila.get(col_nombre, "") or "").strip()
            correo = str(fila.get(col_email, "") or "").strip()
            try:
                cantidad = int(fila.get(col_cantidad, 1) or 1)
            except Exception:
                cantidad = 1

            if not correo or "@" not in correo:
                resumen_items.append({"email": correo, "estado": "invalido", "cartones": []})
                progreso_p = PROGRESOS.get(proc_id, {})
                progreso_p.get("items", []).append({"email": correo, "estado": "invalido", "cartones": []})
                continue

            correo_lower = correo.lower()
            cartones_asignados = []

            if correo_lower in mapa_excepciones:
                rini, rfin = mapa_excepciones[correo_lower]
                cartones_asignados = [n for n in range(rini, rfin + 1) if n in mapa_cartones and n not in cartones_ya_usados]
            else:
                while len(cartones_asignados) < cantidad and pool_idx < len(pool_global):
                    num = pool_global[pool_idx]
                    pool_idx += 1
                    if num not in cartones_ya_usados:
                        cartones_asignados.append(num)

            if not cartones_asignados:
                progreso_p = PROGRESOS.get(proc_id, {})
                progreso_p.get("items", []).append({"email": correo, "estado": "sin_cartones", "cartones": []})
                continue

            for num in cartones_asignados:
                cartones_ya_usados.add(num)

            links_carton = [mapa_cartones.get(n, "") for n in cartones_asignados]
            filas_tabla = ""
            for num, lk in zip(cartones_asignados, links_carton):
                if lk:
                    filas_tabla += f'<tr><td style="padding:8px 14px;border:1px solid #ddd;">Cartón #{num}</td><td style="padding:8px 14px;border:1px solid #ddd;"><a href="{lk}" style="color:#1a56db">Ver cartón</a></td></tr>'
                else:
                    filas_tabla += f'<tr><td style="padding:8px 14px;border:1px solid #ddd;">Cartón #{num}</td><td style="padding:8px 14px;border:1px solid #ddd;">Sin enlace</td></tr>'
            tabla_html = (
                f'<table style="border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px;margin-top:8px;">'
                f'<thead><tr>'
                f'<th style="padding:8px 14px;background:#4F81BD;color:#fff;border:1px solid #ddd;">Cartón</th>'
                f'<th style="padding:8px 14px;background:#4F81BD;color:#fff;border:1px solid #ddd;">Enlace</th>'
                f'</tr></thead><tbody>{filas_tabla}</tbody></table>'
            )

            cuerpo = html_body or "<p>Hola {nombre},<br>Te enviamos tus cartones de bingo.</p><p>{links}</p>"
            try:
                mensaje_html = cuerpo.format(nombre=nombre, email=correo, links=tabla_html)
            except Exception:
                try:
                    mensaje_html = cuerpo.replace("{nombre}", nombre).replace("{links}", tabla_html)
                except Exception:
                    mensaje_html = cuerpo

            if "{links}" not in (html_body or ""):
                mensaje_html = mensaje_html + "<br>" + tabla_html

            if firma:
                mensaje_html += firma

            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email.utils import make_msgid
            import smtplib

            msg = MIMEMultipart("mixed")
            rel = MIMEMultipart("related")
            alt = MIMEMultipart("alternative")
            rel.attach(alt)
            alt.attach(MIMEText(mensaje_html, "html"))
            msg.attach(rel)

            remitente = cuenta_actual["user"] if cuenta_actual else ""
            msg["From"] = remitente
            msg["To"] = correo
            msg["Subject"] = asunto
            msg["Message-ID"] = make_msgid()

            estado = "fallo"
            if not servidor:
                servidor, cuenta_actual = conectar_servidor()

            if cuenta_actual and cuenta_actual.get("daily_count", 0) >= 450:
                from backend.servicio_envios import conectar_servidor as cs
                servidor, cuenta_actual = cs(servidor)

            if servidor and cuenta_actual:
                msg.replace_header("From", cuenta_actual["user"])
                try:
                    import time as _time
                    resp = servidor.sendmail(cuenta_actual["user"], correo, msg.as_string())
                    if isinstance(resp, dict) and resp:
                        estado = "rebotado"
                    else:
                        estado = "enviado"
                        incrementar_contador(cuenta_actual["user"])
                        cuenta_actual["daily_count"] = cuenta_actual.get("daily_count", 0) + 1
                        cuenta_actual["last_sent_date"] = pd.Timestamp.now().isoformat()
                except smtplib.SMTPServerDisconnected:
                    from backend.servicio_envios import conectar_servidor as cs
                    servidor, cuenta_actual = cs(servidor)
                    try:
                        if servidor and cuenta_actual:
                            resp = servidor.sendmail(cuenta_actual["user"], correo, msg.as_string())
                            estado = "enviado" if not (isinstance(resp, dict) and resp) else "rebotado"
                            if estado == "enviado":
                                incrementar_contador(cuenta_actual["user"])
                                cuenta_actual["daily_count"] = cuenta_actual.get("daily_count", 0) + 1
                    except Exception:
                        estado = "fallo"
                except Exception:
                    estado = "fallo"

            prog = PROGRESOS.get(proc_id, {})
            prog.get("items", []).append({
                "email": correo,
                "estado": estado,
                "cartones": cartones_asignados
            })
            prog["status"] = {
                "cuenta_user": cuenta_actual.get("user", "") if cuenta_actual else "",
                "cuenta_daily": cuenta_actual.get("daily_count", 0) if cuenta_actual else 0,
                "cuenta_limit": 450,
                "email_actual": correo
            }

            if estado == "enviado":
                fecha_ahora = datetime.now().isoformat()
                for num in cartones_asignados:
                    envios_realizados.append({
                        "carton": num,
                        "email": correo,
                        "nombre": nombre,
                        "fecha": fecha_ahora
                    })

            import time as _time2
            _time2.sleep(5)
            if len(prog.get("items", [])) % 50 == 0:
                _time2.sleep(30)

        if envios_realizados:
            registrar_envios(drive_url, envios_realizados)

        ruta_rep = os.path.join(app.config["OUTPUT_FOLDER"], "reporte_cartones.xlsx")
        try:
            reporte_completo = obtener_reporte_evento(drive_url)
            from openpyxl import Workbook as WB
            from openpyxl.styles import Font as FT, PatternFill as PF, Alignment as AL, Border as BR, Side as SD
            wb2 = WB()
            ws2 = wb2.active
            ws2.title = "Cartones Enviados"
            encabezados = ["Cartón #", "Nombre", "Email", "Fecha"]
            ws2.append(encabezados)
            fuente_enc = FT(bold=True, color="FFFFFF")
            relleno_enc = PF("solid", fgColor="4F81BD")
            borde = BR(left=SD(style="thin", color="BFBFBF"), right=SD(style="thin", color="BFBFBF"),
                       top=SD(style="thin", color="BFBFBF"), bottom=SD(style="thin", color="BFBFBF"))
            al_c = AL(horizontal="center", vertical="center")
            for c in range(1, 5):
                cel = ws2.cell(row=1, column=c)
                cel.font = fuente_enc
                cel.fill = relleno_enc
                cel.alignment = al_c
                cel.border = borde
            for r_item in reporte_completo:
                ws2.append([r_item["carton"], r_item["nombre"], r_item["email"], r_item["fecha"]])
            for i, _ in enumerate(encabezados, start=1):
                ws2.column_dimensions[__import__('openpyxl').utils.get_column_letter(i)].width = 25
            ws2.freeze_panes = "A2"
            ws3 = wb2.create_sheet("Esta sesión")
            ws3.append(["Cartón #", "Nombre", "Email", "Estado"])
            for c in range(1, 5):
                cel = ws3.cell(row=1, column=c)
                cel.font = fuente_enc
                cel.fill = relleno_enc
                cel.alignment = al_c
                cel.border = borde
            for it in PROGRESOS.get(proc_id, {}).get("items", []):
                for num in it.get("cartones", []):
                    ws3.append([num, "", it.get("email", ""), it.get("estado", "")])
            for i in range(1, 5):
                ws3.column_dimensions[__import__('openpyxl').utils.get_column_letter(i)].width = 25
            wb2.save(ruta_rep)
        except Exception as e:
            logger.error(f"Error generando reporte cartones: {e}")

        PROGRESOS[proc_id]["done"] = True
        try:
            os.remove(ruta_archivo)
        except Exception:
            pass
        if servidor:
            try:
                servidor.quit()
            except Exception:
                pass

    h = threading.Thread(target=tarea, daemon=True)
    h.start()

    registrar_accion(current_user.id, "envio_cartones", {"drive_url": drive_url, "asunto": asunto})
    return jsonify({"ok": True, "id": proc_id})


@app.route("/envio-cartones/progreso", methods=["GET"])
@login_required
def cartones_progreso():
    proc_id = request.args.get("id", "")
    off = int(request.args.get("offset", "0"))
    dat = PROGRESOS.get(proc_id) or {"items": [], "done": True}
    items = dat.get("items", [])
    done = dat.get("done", True)
    status = dat.get("status", {})
    nuevos = items[off:] if off < len(items) else []
    return jsonify({"ok": True, "items": nuevos, "done": done, "total": len(items), "status": status})


@app.route("/envio-cartones/reporte")
@login_required
def cartones_reporte():
    ruta = os.path.join(app.config["OUTPUT_FOLDER"], "reporte_cartones.xlsx")
    if os.path.exists(ruta):
        return send_file(ruta, as_attachment=True, download_name="reporte_cartones.xlsx")
    return "Reporte no disponible", 404


@app.route("/envio-cartones/reporte-evento", methods=["POST"])
@login_required
def cartones_reporte_evento():
    drive_url = request.form.get("drive_url", "").strip()
    if not drive_url:
        return jsonify({"ok": False})
    reporte = obtener_reporte_evento(drive_url)
    usados = len(reporte)
    return jsonify({"ok": True, "total_enviados": usados, "items": reporte[:100]})


if __name__ == "__main__":
    app.run(debug=True, port=5000)
