import os
import re
import sys
from typing import List, Dict

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
ruta_base_drive = os.path.dirname(os.path.abspath(__file__))

def obtener_id_carpeta(url: str) -> str:
    patron = r"/folders/([a-zA-Z0-9_-]+)"
    m = re.search(patron, url)
    if not m:
        raise ValueError("URL de carpeta no válida")
    return m.group(1)

def clave_natural_nombre(x: Dict[str, str]):
    n = str(x.get("nombre") or "").lower()
    partes = re.split(r"(\d+)", n)
    k = []
    for p in partes:
        if p.isdigit():
            k.append(int(p))
        else:
            k.append(p)
    return k

def obtener_nombre_carpeta(servicio, carpeta_id: str) -> str:
    r = servicio.files().get(fileId=carpeta_id, fields="name").execute()
    n = r.get("name") or "Carpeta"
    n = re.sub(r"[<>:\\/\|\?\*]", "_", n)
    return n

def autenticar() -> Credentials:
    creds = None
    ruta_token = os.path.join(ruta_base_drive, "token.json")
    ruta_creds = os.path.join(ruta_base_drive, "credentials.json")
    if os.path.exists(ruta_token):
        creds = Credentials.from_authorized_user_file(ruta_token, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(ruta_creds):
                return None
            flujo = InstalledAppFlow.from_client_secrets_file(ruta_creds, SCOPES)
            creds = flujo.run_local_server(port=0)
        with open(ruta_token, "w", encoding="utf-8") as f:
            f.write(creds.to_json())
    return creds

def listar_archivos(servicio, carpeta_id: str) -> List[Dict[str, str]]:
    datos = []
    pagina = None
    while True:
        res = servicio.files().list(q=f"'{carpeta_id}' in parents and trashed = false", pageSize=1000, pageToken=pagina, fields="nextPageToken, files(id, name, mimeType, webViewLink, size, modifiedTime)").execute()
        archivos = res.get("files", [])
        for a in archivos:
            if a.get("mimeType") == "application/vnd.google-apps.folder":
                continue
            enlace = a.get("webViewLink")
            if not enlace:
                enlace = f"https://drive.google.com/file/d/{a.get('id')}/view"
            tam = 0
            try:
                tam = int(a.get("size") or 0)
            except Exception:
                tam = 0
            datos.append({"nombre": a.get("name"), "enlace": enlace, "mimeType": a.get("mimeType"), "size": tam, "mod": a.get("modifiedTime")})
        pagina = res.get("nextPageToken")
        if not pagina:
            break
    return datos

def guardar_excel(datos: List[Dict[str, str]], ruta_salida: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Enlaces"
    ws.append(["Nombre", "Enlace"])
    for d in datos:
        ws.append([d.get("nombre", ""), d.get("enlace", "")])
    max_col = [0, 0]
    for i, fila in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2), start=2):
        n = fila[0].value or ""
        e = fila[1].value or ""
        max_col[0] = max(max_col[0], len(str(n)))
        max_col[1] = max(max_col[1], len(str(e)))
        c_enlace = ws.cell(row=i, column=2)
        if isinstance(e, str) and e.startswith("http"):
            c_enlace.hyperlink = e
            c_enlace.style = "Hyperlink"
        ws.cell(row=i, column=1).font = Font(bold=False, color="000000")
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    cabecera = [ws.cell(row=1, column=1), ws.cell(row=1, column=2)]
    for c in cabecera:
        c.font = Font(bold=True, color="000000", size=12)
        c.fill = PatternFill("solid", fgColor="87CEEB")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=Side(style="thick", color="87CEEB"))
    ws.row_dimensions[1].height = 24
    ancho_nombre = min(max(90, max_col[0] + 10), 120)
    ancho_enlace = min(max(120, max_col[1] + 6), 150)
    ws.column_dimensions[get_column_letter(1)].width = ancho_nombre
    ws.column_dimensions[get_column_letter(2)].width = ancho_enlace
    if ws.max_row >= 2:
        tabla = Table(displayName="TablaEnlaces", ref=f"A1:B{ws.max_row}")
        estilo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)
    wb.save(ruta_salida)

def main() -> None:
    url = None
    if len(sys.argv) >= 2:
        url = sys.argv[1]
    else:
        ruta_url = os.path.join(ruta_base_drive, "carpeta_url.txt")
        if os.path.exists(ruta_url):
            with open(ruta_url, "r", encoding="utf-8") as f:
                url = f.read().strip()
        else:
            print("Falta argumento URL o archivo carpeta_url.txt")
            sys.exit(1)
    salida = sys.argv[2] if len(sys.argv) > 2 else "enlaces.xlsx"
    carpeta_id = obtener_id_carpeta(url)
    creds = autenticar()
    clave = os.environ.get("GOOGLE_API_KEY")
    if not creds and not clave:
        ruta = os.path.join(ruta_base_drive, "api_key.txt")
        if os.path.exists(ruta):
            with open(ruta, "r", encoding="utf-8") as f:
                clave = f.read().strip()
    if not creds and not clave:
        print("Falta credentials.json, GOOGLE_API_KEY o api_key.txt")
        sys.exit(1)
    servicio = build("drive", "v3", credentials=creds) if creds else build("drive", "v3", developerKey=clave)
    datos = listar_archivos(servicio, carpeta_id)
    datos = sorted(datos, key=lambda x: (clave_natural_nombre(x), str(x.get("enlace") or "")))
    nombre_carpeta = obtener_nombre_carpeta(servicio, carpeta_id)
    dir_salida = os.path.join(ruta_base_drive, nombre_carpeta)
    os.makedirs(dir_salida, exist_ok=True)
    ruta_excel = os.path.join(dir_salida, salida)
    if os.path.exists(ruta_excel):
        t_fecha = datetime.now().strftime("%d_%m_%Y")
        t_hora = datetime.now().strftime("%H_%M_%S")
        base, ext = os.path.splitext(salida)
        ruta_copia = os.path.join(dir_salida, f"{base}_copia_{t_fecha}_{t_hora}{ext}")
        shutil.copy2(ruta_excel, ruta_copia)
    guardar_excel(datos, ruta_excel)
    print(f"Listo: {ruta_excel}")

if __name__ == "__main__":
    main()

